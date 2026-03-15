"""
utils/planilha.py
Lê as planilhas Excel usando openpyxl/xlrd diretamente (sem pandas)
"""

import os
import logging

logger = logging.getLogger(__name__)


def _ler_excel(path: str) -> list[dict]:
    """Lê arquivo Excel e retorna lista de dicts com as linhas"""
    if str(path).endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(path)
        sheet = wb.sheet_by_index(0)
        for name in wb.sheet_names():
            if any(p in name.lower() for p in ["carteira", "portfolio", "teto", "radar", "acoes", "ações"]):
                sheet = wb.sheet_by_name(name)
                break
        headers = [str(sheet.cell_value(0, c)).strip().lower() for c in range(sheet.ncols)]
        rows = []
        for r in range(1, sheet.nrows):
            row = {headers[c]: sheet.cell_value(r, c) for c in range(sheet.ncols)}
            if any(v not in (None, "") for v in row.values()):
                rows.append(row)
        return rows
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet = wb.active
        for name in wb.sheetnames:
            if any(p in name.lower() for p in ["carteira", "portfolio", "teto", "radar", "acoes", "ações"]):
                sheet = wb[name]
                break
        rows_raw = list(sheet.iter_rows(values_only=True))
        if not rows_raw:
            return []
        headers = [str(h).strip().lower() if h else f"col{i}" for i, h in enumerate(rows_raw[0])]
        rows = []
        for row_raw in rows_raw[1:]:
            if all(v is None for v in row_raw):
                continue
            rows.append({headers[i]: row_raw[i] for i in range(len(headers))})
        return rows


def _get(row: dict, alternativas: list, default=None):
    for alt in alternativas:
        if alt.lower() in row and row[alt.lower()] not in (None, ""):
            return row[alt.lower()]
    return default


def _float(val, default=0.0) -> float:
    try:
        return float(val) if val not in (None, "") else default
    except (ValueError, TypeError):
        return default


def _str(val, default="") -> str:
    return str(val).strip() if val is not None else default


def identificar_arquivo(path: str) -> str:
    nome = os.path.basename(path).lower()
    if any(p in nome for p in ["carteira", "portfolio", "posicao"]):
        return "carteira"
    if any(p in nome for p in ["teto", "radar", "acoes", "ações", "monitoramento"]):
        return "radar"
    return "desconhecido"


def carregar_carteira(path: str) -> list[dict]:
    try:
        rows = _ler_excel(path)
        ativos = []
        for row in rows:
            ticker = _str(_get(row, ["ticker", "ativo", "codigo", "código", "papel", "symbol"])).upper()
            if not ticker or ticker in ("NAN", "NONE", ""):
                continue
            # Quantidade pode vir como string com ponto de milhar
            qtd_raw = _get(row, ["quantidade", "qtd", "qtde", "cotas", "shares"])
            qtd = _float(str(qtd_raw).replace(".", "").replace(",", ".") if qtd_raw else 0)

            # Preço médio pode vir com R$, pontos e vírgulas
            pm_raw = _get(row, ["preco_medio", "preço_médio", "preço médio", "pm", "preco medio", "custo medio", "preço médio"])
            pm_str = str(pm_raw).replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".") if pm_raw else "0"
            pm = _float(pm_str)

            mercado = _str(_get(row, ["mercado", "bolsa", "market"], "")).upper()
            if not mercado:
                mercado = _detectar_mercado(ticker)
            classe = _str(_get(row, ["classe", "tipo", "categoria", "setor"], "")).upper()
            if not classe:
                classe = _detectar_classe(ticker)

            ativos.append({
                "ticker":      ticker,
                "quantidade":  qtd,
                "preco_medio": pm,
                "mercado":     mercado,
                "classe":      classe,
            })
        logger.info(f"Carteira carregada: {len(ativos)} ativos")
        return ativos
    except Exception as e:
        logger.error(f"Erro ao carregar carteira: {e}")
        return []


def carregar_radar(path: str) -> list[dict]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)

        # Tenta encontrar aba correta
        aba_alvo = wb.active
        for name in wb.sheetnames:
            if any(p in name.lower() for p in ["dados", "teto", "radar", "acoes", "ações"]):
                aba_alvo = wb[name]
                break

        rows_raw = list(aba_alvo.iter_rows(values_only=True))
        if not rows_raw:
            return []

        # Encontra a linha de cabeçalho (ignora linhas de intro/link)
        header_row = None
        header_idx = 0
        for i, row in enumerate(rows_raw):
            row_str = [str(v).lower() if v else "" for v in row]
            if any(p in " ".join(row_str) for p in ["código", "codigo", "ticker", "ativo"]):
                header_row = row
                header_idx = i
                break

        if not header_row:
            logger.error("Cabeçalho não encontrado na planilha de radar")
            return []

        # Mapeia índices das colunas importantes
        headers = [str(h).strip().lower() if h else "" for h in header_row]

        def find_col(alternativas):
            for alt in alternativas:
                for i, h in enumerate(headers):
                    if alt.lower() in h:
                        return i
            return None

        idx_ticker  = find_col(["código", "codigo", "ticker", "ativo"])
        idx_teto    = find_col(["preço teto", "preco teto", "teto"])
        idx_dy      = find_col(["dividend yield bruto", "dy bruto", "dividend yield"])
        idx_setor   = find_col(["atuação", "atuacao", "setor"])

        if idx_ticker is None:
            logger.error("Coluna de ticker não encontrada")
            return []

        ativos = []
        for row in rows_raw[header_idx+1:]:
            if not row or all(v is None for v in row):
                continue
            ticker = _str(row[idx_ticker] if idx_ticker < len(row) else None).upper()
            if not ticker or len(ticker) < 3 or ticker in ("NAN", "NONE", ""):
                continue
            # Ignora linhas que não são tickers válidos
            if any(p in ticker.lower() for p in ["empresa", "link", "vídeo", "video", "planilha"]):
                continue

            teto = _float(row[idx_teto] if idx_teto and idx_teto < len(row) else None)
            dy   = _float(row[idx_dy]   if idx_dy   and idx_dy   < len(row) else None)
            setor = _str(row[idx_setor] if idx_setor and idx_setor < len(row) else None)

            ativos.append({
                "ticker":     ticker,
                "preco_teto": teto,
                "dy_minimo":  0,
                "pvp_maximo": 0,
                "setor":      setor,
                "notas":      f"DY estimado: {dy:.1%}" if dy else "",
            })

        logger.info(f"Radar carregado: {len(ativos)} ativos")
        return ativos
    except Exception as e:
        logger.error(f"Erro ao carregar radar: {e}")
        return []


def cruzar_carteira_com_teto(carteira: list[dict], radar: list[dict]) -> list[dict]:
    indice_radar = {a["ticker"]: a for a in radar}
    for ativo in carteira:
        dados = indice_radar.get(ativo["ticker"], {})
        ativo.update({
            "preco_teto": dados.get("preco_teto", 0),
            "dy_minimo":  dados.get("dy_minimo", 0),
            "pvp_maximo": dados.get("pvp_maximo", 0),
            "setor":      dados.get("setor", ""),
            "notas":      dados.get("notas", ""),
            "tem_teto":   dados.get("preco_teto", 0) > 0,
        })
    return carteira


def carregar_planilhas(paths: list[str]) -> tuple[list[dict], list[dict]]:
    carteira_path = None
    radar_path    = None
    for path in paths:
        tipo = identificar_arquivo(path)
        if tipo == "carteira":
            carteira_path = path
        elif tipo == "radar":
            radar_path = path

    carteira = carregar_carteira(carteira_path) if carteira_path else []
    radar    = carregar_radar(radar_path)       if radar_path    else []

    if carteira and radar:
        carteira = cruzar_carteira_com_teto(carteira, radar)
    elif carteira:
        for ativo in carteira:
            ativo.update({"preco_teto": 0, "dy_minimo": 0, "pvp_maximo": 0,
                          "setor": "", "notas": "", "tem_teto": False})
    return carteira, radar


def _detectar_mercado(ticker: str) -> str:
    import re
    return "B3" if re.match(r"^[A-Z]{3,4}\d{1,2}$", ticker) else "USA"


def _detectar_classe(ticker: str) -> str:
    import re
    if re.match(r"^[A-Z]{4}11$", ticker):
        return "FII"
    if re.match(r"^[A-Z]{3,4}\d$", ticker):
        return "AÇÃO"
    return "STOCK/ETF"
